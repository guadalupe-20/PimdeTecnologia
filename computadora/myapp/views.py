from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.views.generic.base import TemplateView
from .forms import *
from .models import *
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from django.http.response import HttpResponse
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas

#vista de login 
def home(request):
    if request.user.is_authenticated: # si el usuario es auntenticado 
        return HttpResponseRedirect(reverse('home'))# redirigue al login del sistema 
    login_form = MyAuthForm()#muestra el formulario del login 
    user_form = MyUserCreationForm()#muestra el usuario
    ctx = {
        'login_form': login_form,
        'User_form': user_form
    }
    return render(request, 'home.html',ctx)# me redirigue a la pantalla de login 

#vista para salir del sistema 
def salir(request):
    logout(request)#vista log-out es para poder regresar a la pantalla de login 
    login_form = MyAuthForm()
    user_form = MyUserCreationForm()
    ctx = {
        'login_form': login_form,
        'User_form': user_form
    }
    return render(request, 'home.html',ctx)#me redirige al login 

#vista para acceder al menu del sistema 
def menu(request):
    home(request)
    return render(request,'menu.html')# me redirige al menu 

#vista para el boton que muestra el listado de los reportes 
def reportetotal(request):
    home(request)
    return render(request,'reportetotal.html')# me redirigue a la dirección de la plantilla 

#vista que muestra el listado de los reportes 
def trasladoequipos(request):
    home(request)
    return render(request,'trasladoequipos.html')# me redirigue a la dirección de la plantilla y acceder acada uno de ellos 

#vista para procesar el login con sus repectivo filtro 
def procesar_login(request):
    if request.method == 'POST':# se obtiene el usuario
        form = MyAuthForm(data = request.POST)
        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(username = request.POST['username'], password = request.POST['password'])
            if user is not None:# si el id es Verdadero continua con el acceso al sistema 
                if user.is_active:
                    login(request, user)
                    return HttpResponseRedirect(reverse('menu'))#lo deja acceder a lapantalla del menu 
                else:
                      return HttpResponseRedirect(reverse('Usuario o contraseña incorrecta'))
        else:
             return HttpResponseRedirect(reverse('home'))
    else:
        return HttpResponseRedirect(reverse('No existe esta pagina'))

#vista para la  plantilla de oficina 
@login_required
def oficina(request):
    form = OficinaForm()# se carga el formulario 
    oficina = Oficina.objects.all()#consulta de laoficina 
    ctx = {
        'form': form,#cargar el formulario 
        'oficinas': oficina
    }
    return render(request, 'oficina.html', ctx)#me dedirecciona al fomulario de oficina 

#vista del listado de las ofinas 
@login_required
def oficinalista(request):
    oficina = Oficina.objects.all()# se hace uan consulta a la base de datos 
    queryset = request.GET.get('buscar')# se crea el filtro de la conuslta 
    print(queryset)
    if queryset:#si es correcta la consulta 
        try:
            oficina = Oficina.objects.get(oficina=queryset)#me deja acceder a los datos del filtro 
        except Exception as e:
            oficina = None
            oficina = Oficina.objects.filter(# filto de la consulta 
            Q(codoficina__contains = queryset ) |
            Q(oficina__contains = queryset) 
            ).distinct()
    ctx = {
        'oficinas': oficina# muestra el listdo de las oficinas 
    }
    return render(request, 'oficinalista.html',ctx)# muestra ya la lista de las oficinas 

#vista para almacenar una oficina ingresada 
@login_required
def guardar_oficina(request, id = None):# se toma el id para poder guardar una oficina 
    if request.method == 'POST':
        oficina = get_object_or_404(Oficina, pk=id) if id else None#consulta para aquiri el id
        form = OficinaForm(request.POST or None, instance = oficina)#se carga el formulario 
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('oficina'))
        else:
            return render(request, 'oficina.html', {'form': form})

#vista pa ra eliminar una oficina en especifico  
@login_required
def eliminar_oficina(request, id):
    Oficina.objects.get(pk=id).delete()# se obtiene el id del que se desea eliminar 
    return HttpResponseRedirect(reverse('oficinalista'))# me regresa a la lista de las oficinas 

#bVista para eliminar una oficina en especifico 
@login_required
def editar_oficina(request, id):
    # se crea el objeto oficina
    oficina = Oficina.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto departamento 
    form = OficinaForm(instance = oficina)
    # consulta a todos las odicinas  (para que se sigan viendo al momento de editar)
    oficinas = Oficina.objects.all()
    # se envia el id del empleado para poder marcar con CSS la fila que se esta editando
    return render(request, 'oficina.html', {'form': form, 'oficinas': oficinas , 'idv': id})

#vista para la  plantilla del departamento  
@login_required
def departamento(request):
    form = DepartamentoForm()# se carga el formulario 
    departamento = Departamento.objects.all()#consulta del departamento 
    ctx = {
        'form': form,#cargar el formulario
        'departamentos': departamento
    }
    return render(request, 'departamento.html', ctx)#me dedirecciona al fomulario de departamento

# se crea vista para el listado de los departamentos 
@login_required
def departamentolista(request):
    departamento = Departamento.objects.all()#se crea una consulta para accesar a la informacion del departamento 
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            departamento = Departamento.objects.get(departamento=queryset)# se Filtra la informacion a obtener 
        except Exception as e:
            departamento = None
            departamento = Departamento.objects.filter(#Filtro de la consulta 
            Q(departamento__contains = queryset ) 
            ).distinct()
    ctx = { 
        'departamentos': departamento
    }
    return render(request, 'departamentolista.html', ctx)

# vista para guaradra un nuevo departamento 
@login_required
def guardar_departamento(request, id = None):
    if request.method == 'POST':
        departamento = get_object_or_404(Departamento, pk=id) if id else None
        form = DepartamentoForm(request.POST or None, instance = departamento)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('departamento'))
        else:
            return render(request, 'departamento.html', {'form': form})

# vista ára elimibar un departamento por si id 
@login_required
def eliminar_departamento(request, id):
    Departamento.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('departamentolista'))

@login_required
def editar_departamento(request, id):
    # se crea el objeto departamento
    departamento = Departamento.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto departamento
    form = DepartamentoForm(instance = departamento)
    # consulta a todos los departamento (para que se sigan viendo al momento de editar)
    departamentos = Departamento.objects.all()
    # se envia el id del departamento  para poder marcar con CSS la fila que se esta editando
    return render(request, 'departamento.html', {'form': form, 'departamentos': departamentos , 'idv': id})


@login_required
def marca(request):
    form = MarcaForm()
    marca = Marca.objects.all()
    ctx = {
        'form': form,
        'marcas': marca
    }
    return render(request, 'marca.html', ctx)

@login_required
def marcalista(request):
    marca = Marca.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            marca = Marca.objects.get(marca=queryset)
        except Exception as e:
            marca = None
            marca = Marca.objects.filter(
            Q(marca__contains = queryset )    
             ).distinct()
    ctx = {        
        'marcas': marca
    }
    return render(request, 'marcalista.html', ctx)

@login_required
def guardar_marca(request, id = None):
    if request.method == 'POST':
        marca = get_object_or_404(Marca, pk=id) if id else None
        form = MarcaForm(request.POST or None, instance = marca)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('marca'))
        else:
            return render(request, 'marca.html', {'form': form})

@login_required
def eliminar_marca(request, id):
    Marca.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('marcalista'))

@login_required
def editar_marca(request, id):
    # se crea el objeto marca
    marca = Marca.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto marca
    form = MarcaForm(instance = marca)
    # consulta a todos los marca (para que se sigan viendo al momento de editar)
    marcas = Marca.objects.all()
    # se envia el id del marca para poder marcar con CSS la fila que se esta editando
    return render(request, 'marca.html', {'form': form, 'marcas': marcas , 'idv': id})

@login_required
def modelo(request):
    form = ModeloForm()
    modelo = Modelo.objects.all()
    ctx = {
        'form': form,
        'modelos': modelo
    }
    return render(request, 'modelo.html', ctx)

@login_required
def modelolista(request):
    modelo = Modelo.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            modelo = Modelo.objects.get(modelo=queryset)
        except Exception as e:
            modelo = None
            modelo = Modelo.objects.filter(
            Q(modelo__contains = queryset )            
             ).distinct()
    ctx = {
        'modelos': modelo
    }
    return render(request, 'modelolista.html', ctx)


@login_required
def guardar_modelo(request, id = None):
    if request.method == 'POST':
        modelo = get_object_or_404(Modelo, pk=id) if id else None
        form = ModeloForm(request.POST or None, instance = modelo)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('modelo'))
        else:
            return render(request, 'modelo.html', {'form': form})

@login_required
def eliminar_modelo(request, id):
    Modelo.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('modelolista'))

@login_required
def editar_modelo(request, id):
    # se crea el objeto modelo
    modelo = Modelo.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto modelo
    form = ModeloForm(instance = modelo)
    # consulta a todos los modelo (para que se sigan viendo al momento de editar)
    modelos = Modelo.objects.all()
    # se envia el id del modelo para poder marcar con CSS la fila que se esta editando
    return render(request, 'modelo.html', {'form': form, 'modelos': modelos , 'idv': id})


@login_required
def estado(request):
    form = EstadoForm()
    estado = Estado.objects.all()
    ctx = {
        'form': form,
        'estados': estado
    }
    return render(request, 'estado.html', ctx)

@login_required
def estadolista(request):
    estado = Estado.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            estado = Estado.objects.get(estado=queryset)
        except Exception as e:
            estado = None 
            estado = Estado.objects.filter(
            Q(estado__contains = queryset ) 
            ).distinct()
    ctx = {
        'estados': estado
    }
    return render(request, 'estadolista.html', ctx)

@login_required
def guardar_estado(request, id = None):
    if request.method == 'POST':
        estado = get_object_or_404(Estado, pk=id) if id else None
        form = EstadoForm(request.POST or None, instance = estado)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('estado'))
        else:
            return render(request, 'estado.html', {'form': form})

@login_required
def eliminar_estado(request, id):
    Estado.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('estadolista'))

@login_required
def editar_estado(request, id):
    # se crea el objeto estadolugar reparacion
    estado = Estado.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto estadolugar reparacion
    form = EstadoForm(instance = estado)
    # consulta a todos los estadolugar reparacion (para que se sigan viendo al momento de editar)
    estados = Estado.objects.all()
    # se envia el id del estadolugar reparacion para poder marcar con CSS la fila que se esta editando
    return render(request, 'estado.html', {'form': form, 'estados': estados , 'idv': id})


@login_required
def lugarreparacion(request):
    form = LugarReparacionForm()
    lugarreparacion = LugarReparacion.objects.all()
    ctx = {
        'form': form,
        'lugarreparacions': lugarreparacion
    }
    return render(request, 'lugarreparacion.html', ctx)

@login_required
def lugarreparacionlista(request):
    lugarreparacion = LugarReparacion.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            lugarreparacion = LugarReparacion.objects.get(lugarreparacion=queryset)
        except Exception as e:
            lugarreparacion = None 
            lugarreparacion = LugarReparacion.objects.filter(
            Q(lugarreparacion__contains = queryset )            
             ).distinct()

    ctx = {
        'lugarreparacions': lugarreparacion
    }
    return render(request, 'lugarreparacionlista.html', ctx)

@login_required
def guardar_lugarreparacion(request, id = None):
    if request.method == 'POST':
        lugarreparacion = get_object_or_404(LugarReparacion, pk=id) if id else None
        form = LugarReparacionForm(request.POST or None, instance = lugarreparacion)

        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('lugarreparacion'))
        else:
            return render(request, 'lugarreparacion.html', {'form': form})

@login_required
def eliminar_lugarreparacion(request, id):
    LugarReparacion.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('lugarreparacion'))

@login_required
def editar_lugarreparacion(request, id):
    # se crea el objeto lugar reparacion
    lugarreparacion = LugarReparacion.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto lugar reparacion
    form = LugarReparacionForm(instance = lugarreparacion)
    # consulta a todos los lugar reparacion (para que se sigan viendo al momento de editar)
    lugarreparacions = LugarReparacion.objects.all()
  # se envia el id del lugar reparacion para poder marcar con CSS la fila que se esta editando
    return render(request, 'lugarreparacion.html', {'form': form, 'lugarreparacions': lugarreparacions , 'idv': id})

@login_required

def usuario(request):
    form = UsuarioForm()
    usuario = Usuario.objects.all()
    ctx = {
        'form': form,
        'usuarios': usuario
    }

    return render(request, 'usuario.html', ctx)

@login_required
def usuariolista(request):
    usuario = Usuario.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)

    if queryset:
        try:
            usuario = Usuario.objects.get(usuario=queryset)
        except Exception as e:
            usuario = None
   
            usuario = Usuario.objects.filter(
            Q(codusuario__contains = queryset ) |
            Q(nombre__contains = queryset )|
            Q(apellido__contains = queryset )|
            Q(nombreusuario__contains = queryset )
             ).distinct()

    ctx = {
            'usuarios': usuario
    }
    return render(request, 'usuariolista.html', ctx)

@login_required
def guardar_usuario(request, id = None):
    if request.method == 'POST':
        usuario = get_object_or_404(Usuario, pk=id) if id else None
        form = UsuarioForm(request.POST or None, instance = usuario)

        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('usuario'))
        else:
            return render(request, 'usuario.html', {'form': form})

@login_required
def eliminar_usuario(request, id):
    Usuario.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('usuariolista'))

@login_required
def editar_usuario(request, id):
    # se crea el objeto usuario
    usuario = Usuario.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto usuario
    form = UsuarioForm(instance = usuario)
    # consulta a todos los usuario (para que se sigan viendo al momento de editar)
    usuarios = Usuario.objects.all()
    # se envia el id del usuario para poder marcar con CSS la fila que se esta editando
    return render(request, 'usuario.html', {'form': form, 'usuarios': usuarios , 'idv': id})

@login_required

def auditoria(request):
    form = AuditoriaForm()
    auditoria = Auditoria.objects.all()
    ctx = {
        'form': form,
        'auditorias': auditoria
    }
    return render(request, 'auditoria.html', ctx)


@login_required
def auditorialista(request):
    auditoria = Auditoria.objects.all()
    ctx = {
        'auditorias': auditoria
    }
    return render(request, 'auditorialista.html', ctx)

@login_required
def guardar_auditoria(request, id = None):
    if request.method == 'POST':
        auditoria = get_object_or_404(Auditoria, pk=id) if id else None
        form = AuditoriaForm(request.POST or None, instance = auditoria)

        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('auditoria'))
        else:
            return render(request, 'auditoria.html', {'form': form})


@login_required
def eliminar_auditoria(request, id):
    Auditoria.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('auditorialista'))

@login_required
def editar_auditoria(request, id):
    # se crea el objeto auditoria
    auditoria = Auditoria.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto auditoria
    form = AuditoriaForm(instance = auditoria)
    # consulta a todos los auditoria (para que se sigan viendo al momento de editar)
    auditorias = Auditoria.objects.all()
    # se envia el id del auditoria para poder marcar con CSS la fila que se esta editando
    return render(request, 'auditoria.html', {'form': form, 'auditorias': auditorias , 'idv': id})


@login_required
def procesador(request):
    form = ProcesadorForm()
    procesador = Procesador.objects.all()
    ctx = {
        'form': form,
        'procesadors': procesador
    }
    return render(request, 'procesador.html', ctx)

@login_required
def procesadorlista(request):
    procesador = Procesador.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)

    if queryset:
        try:
            procesador = Procesador.objects.get(procesador=queryset)
        except Exception as e:
            procesador = None
   
            procesador = Procesador.objects.filter(
            Q(procesador__contains = queryset ) 
             ).distinct()

    ctx = {
        'procesadors': procesador
    }
    return render(request, 'procesadorlista.html', ctx)

@login_required
def guardar_procesador(request, id = None):
    if request.method == 'POST':
        procesador = get_object_or_404(Procesador, pk=id) if id else None
        form = ProcesadorForm(request.POST or None, instance = procesador)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('procesador'))
        else:
            return render(request, 'procesador.html', {'form': form})

@login_required
def eliminar_procesador(request, id):
    Procesador.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('procesadorlista'))

@login_required
def editar_procesador(request, id):
    # se crea el objeto procesador
    procesador = Procesador.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto procesador
    form = ProcesadorForm(instance = procesador)
    # consulta a todos los procesador (para que se sigan viendo al momento de editar)
    procesadors = Procesador.objects.all()
    # se envia el id del procesador para poder marcar con CSS la fila que se esta editando
    return render(request, 'procesador.html', {'form': form, 'procesadors': procesadors , 'idv': id})

@login_required
def sistemaoperativo(request):
    form = SistemaOperativoForm()

    sistemaoperativo = SistemaOperativo.objects.all()
    ctx = {
        'form': form,
        'sistemaoperativos': sistemaoperativo
    }
    return render(request, 'sistemaoperativo.html', ctx)

@login_required
def sistemaoperativolista(request):
    sistemaoperativo = SistemaOperativo.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            sistemaoperativo = SistemaOperativo.objects.get(sistemaoperativo=queryset)
        except Exception as e:
            sistemaoperativo = None
   
            sistemaoperativo = SistemaOperativo.objects.filter(
            Q(sistemaoperativo__contains = queryset ) 
             ).distinct()

    ctx = {
        'sistemaoperativos': sistemaoperativo
    }
    return render(request, 'sistemaoperativolista.html', ctx)

@login_required
def guardar_sistemaoperativo(request, id = None):
    if request.method == 'POST':
        sistemaoperativo = get_object_or_404(SistemaOperativo, pk=id) if id else None
        form = SistemaOperativoForm(request.POST or None, instance = sistemaoperativo)

        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('sistemaoperativo'))
        else:
            return render(request, 'sistemaoperativo.html', {'form': form})

@login_required
def eliminar_sistemaoperativo(request, id):
    SistemaOperativo.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('sistemaoperativolista'))

@login_required
def editar_sistemaoperativo(request, id):
    # se crea el objeto sistema operativo
    sistemaoperativo = SistemaOperativo.objects.get(pk=id)
# se crea y se carga el formulario con los datos del objeto sistema operativo
    form = SistemaOperativoForm(instance = sistemaoperativo)
    # consulta a todos los sistema operativo (para que se sigan viendo al momento de editar)
    sistemaoperativos = SistemaOperativo.objects.all()
    # se envia el id del sistema operativo para poder marcar con CSS la fila que se esta editando
    return render(request, 'sistemaoperativo.html', {'form': form, 'sistemaoperativos': sistemaoperativos , 'idv': id})

@login_required
def equipo(request):
    form = EquipoForm()
    equipo = Equipo.objects.all()
    ctx = {
        'form': form,
        'equipos': equipo
    }
    return render(request, 'equipo.html', ctx)

@login_required
def equipolista(request):
    equipo = Equipo.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)

    if queryset:
        try:
            equipo = Equipo.objects.get(equipo=queryset)
        except Exception as e:
            equipo = None
            equipo = Equipo.objects.filter(
            Q(codequipo__contains = queryset ) |
            Q(nombreequipo__contains = queryset ) |
            Q(fechacompra__contains = queryset )|
            Q(servicetag__contains = queryset )|
            Q(direccionip__contains = queryset ) 
             ).distinct()
    ctx = {
        'equipos': equipo
    }
    return render(request, 'equipolista.html', ctx)

def reportetraslado(request):
     campo = request.GET.get('campo') 
     equipo = Equipo.objects.filter(codequipo=campo)
     wb = Workbook()
     ws = wb.active 

     ws['B1'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B1'].font = Font(name ='arial-black' , size = 18, bold = True)

     ws['B2'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B2'].font = Font(name ='arial' , size = 14 , bold = True)

     ws['B6'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B6'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                             ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B6'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B7'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B7'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                        ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B7'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B8'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B8'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                               ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B8'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B9'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B9'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                                ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B9'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B10'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B10'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                            ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B10'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B11'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B11'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                            ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B11'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B12'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B12'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                            ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B12'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B13'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B13'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                            ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B13'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B14'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B14'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                            ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B14'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B15'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B15'].border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin")
                            ,top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B15'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B22'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B22'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B27'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B27'].font = Font(name ='arial' , size = 10 , bold = True)
 
     ws.column_dimensions['B'].width =30
     ws.column_dimensions['C'].width =60
     ws.row_dimensions[6].height=25
     ws.row_dimensions[7].height=25
     ws.row_dimensions[8].height=25
     ws.row_dimensions[9].height=25
     ws.row_dimensions[10].height=25
     ws.row_dimensions[11].height=25
     ws.row_dimensions[12].height=30
     ws.row_dimensions[13].height=30
     ws.row_dimensions[14].height=25
     ws.row_dimensions[15].height=30
     ws['B1'] = 'Traslado de Equipos de Computo'
     ws['B2'] = 'Cooperativa de Ahorro y Credito La Guadalupe LTDA.'
     ws.merge_cells('B1:D1')
     ws.merge_cells('B2:D2')
     ws.merge_cells('B22:D22')
     ws.merge_cells('B27:D27')
     ws['B6'] = 'FECHA TRASLADO:'
     ws['B7'] = 'NOMBRE DEL EQUIPO:'
     ws['B8'] = 'DESCIRPCIÓN:'
     ws['B9'] = 'NÚMERO DE INVETARIO:'
     ws['B10'] = 'OFICINA/VENTANILLA:'
     ws['B11'] = 'RESPONSABLE:'
     ws['B12'] = 'MOTIVO DEL MOVIMIENTO:'
     ws['B13'] = 'DIAGNOSTICO:'
     ws['B14'] = 'LUGAR DE REPARACIÓN:'
     ws['B15'] = 'OBSERVACIONES:'
     ws['B22'] = 'FIRMA DEL RESPONSABLE:  ____________________________________________________'
     ws['B27'] = 'FIRMA DEL MOVIMIENTO:   ____________________________________________________'
     cont=3
     for  equi in equipo:
        ws.cell(row = 6, column = cont).value = equi.fechacompra
        ws.cell(row = 6, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 6, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 7, column = cont).value =  equi.nombreequipo
        ws.cell(row = 7, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 7, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 8, column = cont).value =  equi.descripcion
        ws.cell(row = 8, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 8, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 9, column = cont).value = equi.codequipo
        ws.cell(row = 9, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 9, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 10, column = cont).value = str(equi.oficina)
        ws.cell(row = 10, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 10, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 11, column = cont).value = str(equi.usuario)
        ws.cell(row = 11, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 11, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 14, column = cont).value = str(equi.lugarreparacion)
        ws.cell(row = 14, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 14, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 12, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 13, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 15, column = cont).border = Border(left = Side(border_style = "thin"), right 
                                 =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        cont = cont
                
        nombre_archivo = "Reporte_Traslado.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response 

def reportediagnostico(request):
     campo = request.GET.get('campo') 
     equipo = Equipo.objects.filter(codequipo=campo)
     wb = Workbook()
     ws = wb.active 

     ws['B1'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B1'].font = Font(name ='arial-black' , size = 18, bold = True)

     ws['B2'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B2'].font = Font(name ='arial' , size = 14 , bold = True)

     ws['B6'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B6'].border = Border(left = Side(border_style = "thin"), right 
                       =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B6'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B7'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B7'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B7'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B8'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B8'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B8'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B9'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B9'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B9'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B10'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B10'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B10'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B11'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B11'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B11'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B12'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B12'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B12'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B13'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B13'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B13'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B14'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B14'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B14'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B15'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B15'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B15'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B16'].alignment = Alignment(horizontal ='right' , vertical='center')
     ws['B16'].border = Border(left = Side(border_style = "thin"), right =
                       Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws['B16'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B22'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B22'].font = Font(name ='arial' , size = 10 , bold = True)
     ws['B27'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B27'].font = Font(name ='arial' , size = 10 , bold = True)

     ws.column_dimensions['B'].width =30
     ws.column_dimensions['C'].width =60
     ws.row_dimensions[6].height=25
     ws.row_dimensions[7].height=25
     ws.row_dimensions[8].height=25
     ws.row_dimensions[9].height=25
     ws.row_dimensions[10].height=25
     ws.row_dimensions[11].height=25
     ws.row_dimensions[12].height=30
     ws.row_dimensions[13].height=30
     ws.row_dimensions[14].height=25
     ws.row_dimensions[15].height=30
     ws.row_dimensions[16].height=25

     ws['B1'] = 'Diagnostico de Equipos de Computo'
     ws['B2'] = 'Cooperativa de Ahorro y Credito La Guadalupe LTDA.'

     ws.merge_cells('B1:D1')
     ws.merge_cells('B2:D2')
     ws.merge_cells('B22:D22')
     ws.merge_cells('B27:D27')
     ws['B6'] = 'FECHA DIAGNOSTICO:'
     ws['B7'] = 'NOMBRE DEL EQUIPO:'
     ws['B8'] = 'DESCIRPCIÓN:'
     ws['B9'] = 'NÚMERO DE INVETARIO:'
     ws['B10'] = 'OFICINA/VENTANILLA:'
     ws['B11'] = 'RESPONSABLE:'
     ws['B12'] = 'MOTIVO DEL DIAGNOSTICO:'
     ws['B13'] = 'COMO SE DETECTO:'
     ws['B14'] = 'DIAGNOSTICADA POR:'
     ws['B15'] = 'OBSERVACIONES:'
     ws['B16'] = 'FECHA COMPRA:'
     ws['B22'] = 'FIRMA DEL RESPONSABLE:  ____________________________________________________'
     ws['B27'] = 'FIRMA DEL MOVIMIENTO:   ____________________________________________________'
     cont=3
     for  equi in equipo:
        ws.cell(row = 6, column = cont).value = equi.fechacompra
        ws.cell(row = 6, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 6, column = cont).border = Border(left = Side(border_style = "thin")
                                 , right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 7, column = cont).value =  equi.nombreequipo
        ws.cell(row = 7, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 7, column = cont).border = Border(left = Side(border_style = "thin")
                                 , right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 8, column = cont).value =  equi.descripcion
        ws.cell(row = 8, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 8, column = cont).border = Border(left = Side(border_style = "thin")
                                 , right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 9, column = cont).value = equi.codequipo
        ws.cell(row = 9, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 9, column = cont).border = Border(left = Side(border_style = "thin")
                                 , right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 10, column = cont).value = str(equi.oficina)
        ws.cell(row = 10, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 10, column = cont).border = Border(left = Side(border_style = "thin")
                                 , right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 11, column = cont).value = str(equi.usuario)
        ws.cell(row = 11, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 11, column = cont).border = Border(left = Side(border_style = "thin")
                                 , right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 14, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 14, column = cont).border = Border(left = Side(border_style = "thin")
                                 , right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 12, column = cont).border = Border(left = Side(border_style = "thin")
                                 , right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 13, column = cont).border = Border(left = Side(border_style = "thin"
                                 ), right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 15, column = cont).border = Border(left = Side(border_style = "thin")
                                 , right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        ws.cell(row = 16, column = cont).value = equi.fechacompra
        ws.cell(row = 16, column = cont).alignment = Alignment(horizontal ='center' , vertical='center')
        ws.cell(row = 16, column = cont).border = Border(left = Side(border_style = "thin"), right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
        cont = cont
                
        nombre_archivo = "Reporte_Diagnostico.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response 

@login_required
def guardar_equipo(request, id = None):
    if request.method == 'POST':
      equipo = get_object_or_404(Equipo, pk=id) if id else None
      form = EquipoForm(request.POST or None, instance = equipo)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('equipo'))
    else:
        return render(request, 'equipo.html', {'form': form})

@login_required
def eliminar_equipo(request, id):
    Equipo.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('equipolista'))

@login_required
def editar_equipo(request, id):
    # se crea el objeto empleado
    equipo = Equipo.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto empleado
    form = EquipoForm(instance = equipo)
    # consulta a todos los empleados (para que se sigan viendo al momento de editar)
    equipos = Equipo.objects.all()
    # se envia el id del empleado para poder marcar con CSS la fila que se esta editando
    return render(request, 'equipo.html', {'form': form, 'equipos': equipos , 'idv': id})
            
@login_required
def mantenimiento(request):
    form = MantenimientoPreventivoForm()
    mantenimientopreventivo = MantenimientoPreventivo.objects.all()
    ctx = {
        'form': form,
        'mantenimientopreventivos': mantenimientopreventivo
    }
    return render(request, 'mantenimiento.html', ctx)

@login_required
def mantenimientolista(request):
    mantenimientopreventivo = MantenimientoPreventivo.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            mantenimientopreventivo = MantenimientoPreventivo.objects.get(mantenimientopreventivo=queryset)
        except Exception as e:
            mantenimientopreventivo = None
            mantenimientopreventivo = MantenimientoPreventivo.objects.filter(
            Q(fechamantenimiento__contains = queryset ) 
             ).distinct()
    ctx = {
        'mantenimientopreventivos': mantenimientopreventivo
    }
    return render(request, 'mantenimientolista.html', ctx)

@login_required
def guardar_mantenimientopreventivo(request, id = None):
    if request.method == 'POST':
      mantenimientopreventivo = get_object_or_404(MantenimientoPreventivo, pk=id) if id else None
      form = MantenimientoPreventivoForm(request.POST or None, instance = mantenimientopreventivo)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('mantenimiento'))
    else:
        return render(request, 'mantenimiento.html', {'form': form})

def reporte(request,*args,**kwargs):
    queryset = request.GET.get('.form-control input-sm')
    print(queryset)
    if queryset :
        mantenimiento = MantenimientoPreventivo.objects.filter(
            Q(fechamantenimiento__contains = queryset )
            
            ).distinct()
    else:
        mantenimiento = MantenimientoPreventivo.objects.all()

    wb = Workbook()
    ws = wb.active 

    ws['B1'].alignment = Alignment(horizontal ='center' , vertical='center')
    ws['B1'].font = Font(name ='arial-black' , size = 16 , bold = True)

    ws['B2'].alignment = Alignment(horizontal ='center' , vertical='center')
    ws['B2'].font = Font(name ='arial' , size = 12 , bold = True)

    ws['B4'].alignment = Alignment(horizontal ='center' , vertical='center')
    ws['B4'].font = Font(name ='arial' , size = 10 , bold = True)
    ws['B4'].fill = PatternFill(start_color='AFDAA0',end_color='AFDAA0',fill_type='solid')
    ws['B4'].border = Border(left = Side(border_style = "thin"), right 
                         =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
    ws.column_dimensions['B'].width =22

    ws['C4'].alignment = Alignment(horizontal ='center' , vertical='center')
    ws['C4'].font = Font(name ='arial' , size = 10 , bold = True)
    ws['C4'].fill = PatternFill(start_color='AFDAA0',end_color='AFDAA0',fill_type='solid')
    ws['C4'].border = Border(left = Side(border_style = "thin"), right
                      =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
    ws.column_dimensions['C'].width =20
 
    ws['D4'].alignment = Alignment(horizontal ='center' , vertical='center')
    ws['D4'].font = Font(name ='arial' , size = 10 , bold = True)
    ws['D4'].fill = PatternFill(start_color='AFDAA0',end_color='AFDAA0',fill_type='solid')
    ws['D4'].border = Border(left = Side(border_style = "thin"), right
                      =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
    ws.column_dimensions['D'].width =50

    ws['E4'].alignment = Alignment(horizontal ='center' , vertical='center')
    ws['E4'].font = Font(name ='arial' , size = 10 , bold = True)
    ws['E4'].fill = PatternFill(start_color='AFDAA0',end_color='AFDAA0',fill_type='solid')
    ws['E4'].border = Border(left = Side(border_style = "thin"), right 
                         =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
    ws.column_dimensions['E'].width =20
    ws['F4'].alignment = Alignment(horizontal ='center' , vertical='center')
    ws['F4'].font = Font(name ='arial' , size = 10 , bold = True)
    ws['F4'].fill = PatternFill(start_color='AFDAA0',end_color='AFDAA0',fill_type='solid')
    ws['F4'].border = Border(left = Side(border_style = "thin"), right
                      =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
    ws.column_dimensions['F'].width =25
    ws['B1'] = 'Reporte Mantenimiento Preventivo'
    ws['B2'] = 'Cooperativa de Ahorro y Credito La Guadalupe LTDA.'

    ws.merge_cells('B1:F1')
    ws.merge_cells('B2:F2')
    ws['B4'] = 'Fecha Mantenimiento'
    ws['C4'] = 'Realizado'
    ws['D4'] = 'Descripción'
    ws['E4'] = 'Oficina/Ventanilla'
    ws['F4'] = 'Usuario'
    cont=5

    for  mantenimiento in mantenimiento:
     ws.cell(row = cont, column = 2).value = mantenimiento.fechamantenimiento
     ws.cell(row = cont, column = 2).alignment = Alignment(horizontal ='center' , vertical='center')
     ws.cell(row = cont, column = 2).border = Border(left = Side(border_style 
                                  = "thin"), right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws.cell(row = cont, column = 3).value = mantenimiento.realizado
     ws.cell(row = cont, column = 3).alignment = Alignment(horizontal ='center' , vertical='center')
     ws.cell(row = cont, column = 3).border = Border(left = Side(border_style 
                                  = "thin"), right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws.cell(row = cont, column = 4).value = mantenimiento.descripcion
     ws.cell(row = cont, column = 4).alignment = Alignment(horizontal ='center' , vertical='center')
     ws.cell(row = cont, column = 4).border = Border(left = Side(border_style 
                                  = "thin"), right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws.cell(row = cont, column = 5).value = str(mantenimiento.oficina)
     ws.cell(row = cont, column = 5).alignment = Alignment(horizontal ='center' , vertical='center')
     ws.cell(row = cont, column = 5).border = Border(left = Side(border_style 
                                  = "thin"), right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     ws.cell(row = cont, column = 6).value = str(mantenimiento.usuario)
     ws.cell(row = cont, column = 6).alignment = Alignment(horizontal ='center' , vertical='center')
     ws.cell(row = cont, column = 6).border = Border(left = Side(border_style 
                                  = "thin"), right =Side(border_style = "thin"),top = Side(border_style = "thin"),bottom=Side(border_style = "thin"))
     cont+=1
               
    nombre_archivo = "Reporte_Mantenimiento.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename={0}".format(nombre_archivo)
    response['Content-Disposition'] = content
    wb.save(response)
    return response 

@login_required
def eliminar_mantenimientopreventivo(request, id):
    MantenimientoPreventivo.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('mantenimientolista'))

@login_required
def editar_mantenimientopreventivo(request, id):
    # se crea el objeto empleado
    mantenimientopreventivo = MantenimientoPreventivo.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto empleado
    form = MantenimientoPreventivoForm(instance = mantenimientopreventivo)
    # consulta a todos los empleados (para que se sigan viendo al momento de editar)
    mantenimientopreventivos = MantenimientoPreventivo.objects.all()
    # se envia el id del empleado para poder marcar con CSS la fila que se esta editando
    return render(request, 'mantenimiento.html', {'form': form, 'mantenimientopreventivos': mantenimientopreventivos , 'idv': id})

@login_required
def airedatacenter(request):
    form = AireDatacenterForm()
    airedatacenter = AireDatacenter.objects.all()
    ctx = {
        'form': form,
        'airedatacenters': airedatacenter
    }
    return render(request, 'airedatacenter.html', ctx)

@login_required
def airedatacenterlista(request):
    airedatacenter = AireDatacenter.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            airedatacenter = AireDatacenter.objects.get(airedatacenter=queryset)
        except Exception as e:
            airedatacenter = None
   
            airedatacenter = AireDatacenter.objects.filter(
            Q(codaire__contains = queryset ) |
            Q(nombreaire__contains = queryset )
           
             ).distinct()
    ctx = {
        'airedatacenters': airedatacenter
    }
    return render(request, 'airedatacenterlista.html', ctx)

@login_required
def guardar_airedatacenter(request, id = None):
    if request.method == 'POST':
      airedatacenter = get_object_or_404(AireDatacenter, pk=id) if id else None
      form = AireDatacenterForm(request.POST or None, instance = airedatacenter)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('airedatacenter'))
      else:
        return render(request, 'airedatacenter.html', {'form': form})
 
@login_required
def eliminar_airedatacenter(request, id):
    AireDatacenter.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('airedatacenterlista'))

@login_required
def editar_airedatacenter(request, id):
    # se crea el objeto empleado
    airedatacenter = AireDatacenter.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto empleado
    form = AireDatacenterForm(instance = airedatacenter)
    # consulta a todos los empleados (para que se sigan viendo al momento de editar)
    airedatacenters = AireDatacenter.objects.all()
    # se envia el id del empleado para poder marcar con CSS la fila que se esta editando
    return render(request, 'airedatacenter.html', {'form': form, 'airedatacenters': airedatacenters , 'idv': id})


# # Create your views here.

@login_required

def protocolo(request):
    form = ProtocoloForm()

    protocolo = Protocolo.objects.all()
    ctx = {
        'form': form,
        'protocolos': protocolo
    }

    return render(request, 'protocolo.html', ctx)

@login_required

def protocololista(request):
    protocolo = Protocolo.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)

    if queryset:
        try:
            protocolo = Protocolo.objects.get(protocolo=queryset)
        except Exception as e:
            protocolo = None
   
            protocolo = Protocolo.objects.filter(
            Q(fechaprotocolo__contains = queryset ) |
            Q(historialeventos__contains = queryset )|
            Q(observaciones__contains = queryset )
           
             ).distinct()

    ctx = {
        'protocolos': protocolo
    }
    return render(request, 'protocololista.html', ctx)


def reporteprotocolo(request):
     campo = request.GET.get('campo') 
     protocolo = Protocolo.objects.filter(fechaprotocolo=campo)
     detalle = DetalleAire.objects.all().select_related('protocolo','aire')
     detalles = DetalleAire.objects.all().select_related('protocolo','aire')
     deups = DetalleLectUps.objects.all().select_related('protocolo','ups')
     alarma = AlarmaSistemas.objects.all().select_related('protocolo')
     dats = DetalleAts.objects.all().select_related('protocolo').filter(ats_id=1)
     datss = DetalleAts.objects.all().select_related('protocolo').filter(ats_id=2)
     dat = DetalleAts.objects.all().select_related('protocolo').filter(ats_id=3)
   
     wb = Workbook()
     ws = wb.active 

     ws['B1'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B1'].font = Font(name ='arial-black' , size = 18, bold = True)

     ws['B2'].alignment = Alignment(horizontal ='center' , vertical='center')
     ws['B2'].font = Font(name ='arial' , size = 14 , bold = True)
 
     ws.column_dimensions['B'].width =15
     ws.column_dimensions['C'].width =20
     ws.column_dimensions['D'].width =12
     ws.column_dimensions['E'].width =20
     ws.column_dimensions['F'].width =13
     ws.column_dimensions['G'].width =20

     ws['B1'] = 'Protocolo de Verificacion Diaria del Datacenter'
     ws['B2'] = 'Cooperativa de Ahorro y Credito La Guadalupe LTDA.'


     ws.merge_cells('B1:H1')
     ws.merge_cells('B2:H2')

     ws['B4'] = 'FECHA:'
     ws['D4'] = 'HORA:'
     ws['F4'] = 'OPERADOR:'
     ws['B6'] = 'ALARMA EN UPS:'
     ws['B8'] = 'ALARMA EN AIRES :'
     ws['B10'] = 'ALARMA EN SISTEMAS:'
     ws['C11'] = 'CONTRA INCENDIO:'
     ws['C12'] = 'HABILITADO:'
     ws['B14'] = 'VERIFICAR HISTORIAL DE EVENTOS:'
     ws['B16'] = 'TEMPERATURA DATA:'
     ws['E16'] = 'HUMEDAD DATA :'
     ws['B18'] = 'EXISTE UN EQUIPO QUE NESECITE REVISIÓN:'
     ws['B20'] = 'OBSERVACIONES:'
     ws['B22'] ='AIRE ENCENDIDO :'
     ws['F22'] ='HORA: '
     ws['B24'] ='TEMPERATURA: '
     ws['D24'] ='HUMEDAD:'
     ws['B26'] ='PANEL NO REGULADO:'
     ws['B28'] ='PANEL REGULADO:'
     ws['B30'] ='CILINDRO PRINCIPAL LLENO :'
     ws['B32'] ='CILINDRO MANUAL EN POSICIÓN CORRECTA:'
     ws['B34'] ='UPS:'
     ws['D34'] ='BATERIA %:'
     ws['B36'] ='CARGA % :'
     ws['D36'] ='VSAL:'
     ws['F36'] ='VENT:'
     ws['B38'] ='HZ :'
     ws['D38'] ='CZAS:' 
     ws['F38'] ='PR.HORA:' 
     ws['H38'] ='PR. MIN :'
     ws['B40'] ='VENTOSAS EN POSICIÓN CORRECTA:'
     ws['B42'] ='BOQUILLA AEREA EN POSICIÓN CORRECTA:'
     ws['B44'] ='BOQUILLA BAJO PISO EN POSICIÓN CORRECTA:'
     ws['B46'] ='LUCES PRINCIPALES EN BUEN ESTADO:'
     ws['B48'] ='Ats:'
     ws['D48'] ='A Y B ENCENDIDAS:'
     ws['G48'] ='PREFERENCIA:'
     ws['I48'] ='AMPS:'
     ws['B50'] ='Ats:'
     ws['D50'] ='A Y B ENCENDIDAS:'
     ws['G50'] ='PREFERENCIA:'
     ws['I50'] ='AMPS:'
     ws['B52'] ='Ats:'
     ws['D52'] ='A Y B ENCENDIDAS:'
     ws['G52'] ='PREFERENCIA:'
     ws['I52'] ='AMPS:'
     ws['B54'] ='CAMARA ENVIANDO FOTOS:'
     ws['B54'] ='CONDENSADORA ENCENDIDA:'
     ws['B56'] ='SUPRESOR DE TRASCIENTES OA VERDE:'
     ws['B58'] ='SUPRESOR DE TRASCIENTES OC VERDE:'
     ws['B60'] ='LUZ DE SALIDA ENCENDIDA:'

     cont=4

     for  proto in protocolo:
        ws.cell(row = 4, column =3).value = proto.fechaprotocolo
        ws.cell(row = 4, column =7).value = str(proto.usuario)
        ws.cell(row = 4, column =5).value = proto.hora
        ws.cell(row = 14, column =4).value = proto.historialeventos
        ws.cell(row = 16, column =4).value = proto.temperaturadata
        ws.cell(row = 16, column =6).value = proto.humedaddata
        ws.cell(row = 18, column =5).value = proto.revisionequipos
        ws.cell(row = 20, column =4).value = proto.observaciones
        ws.cell(row = 26, column =4).value =proto.panelnoregulado 
        ws.cell(row = 28, column =4).value =proto.panelregulado 
        ws.cell(row = 40, column =5).value = proto.ventosasenposicion
        ws.cell(row = 42, column =5).value = proto.boquillaaerea
        ws.cell(row = 44, column =5).value = proto.boquillabajopiso
        ws.cell(row = 46, column =6).value = proto.lucesprincipalesbuenestado
        ws.cell(row = 56, column =4).value = proto.oaverde
        ws.cell(row = 58, column =4).value = proto.ocverde
        ws.cell(row = 60, column =4).value = proto.lucessalida


        for  det in detalle:
            ws.cell(row = 8, column =4).value = det.alarmaaire
            ws.cell(row = 22, column =4).value = str(det.aire) 
            ws.cell(row = 22, column =5).value =det.aireencendido
            ws.cell(row = 22, column =7).value =det.horaaire
            ws.cell(row = 24, column =3).value =det.temperaturaaire
            ws.cell(row = 24, column =5).value =det.humedadaire
            for dups in deups:
               ws.cell(row = 6, column =4).value = dups.alarmaups 
               ws.cell(row = 34, column =3).value = str(dups.ups)
               ws.cell(row = 34, column =5).value = dups.bateriaups  
               ws.cell(row = 36, column =3).value = dups.cargaups 
               ws.cell(row = 36, column =5).value = dups.vsalups
               ws.cell(row = 36, column =7).value = dups.ventups
               ws.cell(row = 38, column =3).value = dups.hertzups
               ws.cell(row = 38, column =5).value = dups.czasups
               ws.cell(row = 38, column =7).value = dups.prhorasups
               ws.cell(row = 38, column =9).value = dups.prminutosups      

               for ala in alarma:
                ws.cell(row = 11, column =4).value = ala.alarmacontraincendio 
                ws.cell(row = 12, column =4).value = ala.alarmacontraincendiohabilitado 
                ws.cell(row = 30, column =4).value =ala.cilindroprincipallleno
                ws.cell(row = 32, column =4).value =ala.cilindromanual

                for at in dats:
                    ws.cell(row = 48, column =3).value = str(at.ats)
                    ws.cell(row = 48, column =6).value = at.lucesencendidasayb
                    ws.cell(row = 48, column =8).value = at.preferencia
                    ws.cell(row = 48, column =10).value = at.amps
                    for at in datss:
                        ws.cell(row = 50, column =3).value = str(at.ats)
                        ws.cell(row = 50, column =6).value = at.lucesencendidasayb
                        ws.cell(row = 50, column =8).value = at.preferencia
                        ws.cell(row = 50, column =10).value = at.amps
                        for at in dat:
                            ws.cell(row = 52, column =3).value = str(at.ats)
                            ws.cell(row = 52, column =6).value = at.lucesencendidasayb
                            ws.cell(row = 52, column =7).value = at.preferencia
                            ws.cell(row = 52, column =10).value = at.amps
                            for  de in detalles:
                                ws.cell(row = 54, column =4).value = de.aireencendido

        cont = cont
                
        nombre_archivo = "Reporte_Protocolo.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response 


@login_required
def guardar_protocolo(request, id = None):
    if request.method == 'POST':
      protocolo = get_object_or_404(Protocolo, pk=id) if id else None
      form = ProtocoloForm(request.POST or None, instance = protocolo)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('protocolo'))
    else:
        return render(request, 'protocolo.html', {'form': form})


@login_required
def eliminar_protocolo(request, id):
    Protocolo.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('protocololista'))

@login_required
def editar_protocolo(request, id):
    # se crea el objeto empleado
    protocolo = Protocolo.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto empleado
    form = ProtocoloForm(instance = protocolo)
    # consulta a todos los empleados (para que se sigan viendo al momento de editar)
    protocolos = Protocolo.objects.all()
    # se envia el id del empleado para poder marcar con CSS la fila que se esta editando
    return render(request, 'protocolo.html', {'form': form, 'protocolos': protocolos , 'idv': id})

@login_required
def detalleaire(request):
    form=DetalleAireForm()
    detalleaire = DetalleAire.objects.all()
    ctx = {
        'form': form,
        'detalleaires': detalleaire
    }
    return render(request, 'detalleaire.html', ctx)

@login_required
def detalleairelista(request):
    detalleaire = DetalleAire.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)

    if queryset:
        try:
            detalleaire = DetalleAire.objects.get(detalleaire=queryset)
        except Exception as e:
            detalleaire = None
            detalleaire = DetalleAire.objects.filter(
            Q(alarmaaire__contains = queryset ) |
            Q(aireencendido__contains = queryset )
             ).distinct()
    ctx = {
       
        'detalleaires': detalleaire
    }
    return render(request, 'detalleairelista.html', ctx)

@login_required
def guardar_detalleaire(request, id = None):
    if request.method == 'POST':
      detalleaire = get_object_or_404(DetalleAire, pk=id) if id else None
      form = DetalleAireForm(request.POST or None, instance = detalleaire)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('detalleaire'))
    else:
        return render(request, 'detalleaire.html', {'form': form})

@login_required
def eliminar_detalleaire(request, id):
    DetalleAire.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('detalleairelista'))

@login_required
def editar_detalleaire(request, id):
    # se crea el objeto empleado
    detalleaire = DetalleAire.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto empleado
    form = DetalleAireForm(instance = detalleaire)
   # consulta a todos los empleados (para que se sigan viendo al momento de editar)
    detalleaires = DetalleAire.objects.all()
    # se envia el id del empleado para poder marcar con CSS la fila que se esta editando
    return render(request, 'detalleaire.html', {'form': form, 'detalleaires': detalleaires , 'idv': id})

@login_required
def ups(request):
    form = UpsForm()
    ups = Ups.objects.all()
    ctx = {
        'form': form,
        'upss': ups
    }
    return render(request, 'ups.html', ctx)

@login_required
def upslista(request):
    ups = Ups.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            ups = Ups.objects.get(ups=queryset)
        except Exception as e:
            ups = None
   
            ups = Ups.objects.filter(
            Q(codups__contains = queryset ) |
            Q(nombreups__contains = queryset )
             ).distinct()
    ctx = {
        'upss': ups
    }
    return render(request, 'upslista.html', ctx)

@login_required
def guardar_ups(request, id = None):
    if request.method == 'POST':
      ups = get_object_or_404(Ups, pk=id) if id else None
      form = UpsForm(request.POST or None, instance = ups)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('ups'))
    else:
        return render(request, 'ups.html', {'form': form})

@login_required
def eliminar_ups(request, id):
    Ups.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('upslista'))

@login_required
def editar_ups(request, id):
    # se crea el objeto empleado
    ups = Ups.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto empleado
    form = UpsForm(instance = ups)
    # consulta a todos los empleados (para que se sigan viendo al momento de editar)
    upss = Ups.objects.all()
    # se envia el id del empleado para poder marcar con CSS la fila que se esta editando
    return render(request, 'ups.html', {'form': form, 'upss': upss , 'idv': id})

@login_required
def detallelectups(request):
    form = DetalleLectUpsForm()
    detallelectups = DetalleLectUps.objects.all()
    ctx = {
        'form': form,
        'detallelectupss': detallelectups
    }
    return render(request, 'detallelectups.html', ctx)

@login_required
def detallelectupslista(request):
    detallelectups = DetalleLectUps.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            detallelectups = DetalleLectUps.objects.get(detallelectups=queryset)
        except Exception as e:
            detallelectups = None
            detallelectups = DetalleLectUps.objects.filter(
            Q(alarmaups__contains = queryset ) |
            Q(bateriaups__contains = queryset )|
            Q(cargaups__contains = queryset )
             ).distinct()
    ctx = {
        'detallelectupss': detallelectups
    }
    return render(request, 'detallelectupslista.html', ctx)

@login_required
def guardar_detallelectups(request, id = None):
    if request.method == 'POST':
      detallelectups = get_object_or_404(DetalleLectUps, pk=id) if id else None
      form = DetalleLectUpsForm(request.POST or None, instance = detallelectups)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('detallelectups'))
    else:
        return render(request, 'detallelectups.html', {'form': form})

@login_required
def eliminar_detallelectups(request, id):
    DetalleLectUps.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('detallelectupslista'))

@login_required
def editar_detallelectups(request, id):
    # se crea el objeto empleado
    detallelectups= DetalleLectUps.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto empleado
    form = DetalleLectUpsForm(instance = detallelectups)
    # consulta a todos los empleados (para que se sigan viendo al momento de editar)
    detallelectupss = DetalleLectUps.objects.all()
    # se envia el id del empleado para poder marcar con CSS la fila que se esta editando
    return render(request, 'detallelectups.html', {'form': form, 'detallelectupss': detallelectupss , 'idv': id})

@login_required
def alarmasistemas(request):
    form = AlarmaSistemasForm()
    alarmasistemas = AlarmaSistemas.objects.all()
    ctx = {
        'form': form,
        'alarmasistemass': alarmasistemas
    }
    return render(request, 'alarmasistemas.html', ctx)

@login_required
def alarmasistemaslista(request):
    alarmasistemas = AlarmaSistemas.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)

    if queryset:
        try:
            alarmasistemas = AlarmaSistemas.objects.get(alarmasistemas=queryset)
        except Exception as e:
            alarmasistemas = None
            alarmasistemas = AlarmaSistemas.objects.filter(
            Q(alarmacontraincendio__contains = queryset ) |
            Q(alarmacontraincendiohabilitado__contains = queryset )|
            Q(cilindroprincipallleno__contains = queryset )
             ).distinct()
    ctx = {
        'alarmasistemass': alarmasistemas
    }
    return render(request, 'alarmasistemaslista.html', ctx)

@login_required
def guardar_alarmasistemas(request, id = None):
    if request.method == 'POST':
      alarmasistemas = get_object_or_404(AlarmaSistemas, pk=id) if id else None
      form = AlarmaSistemasForm(request.POST or None, instance = alarmasistemas)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('alarmasistemas'))
    else:
        return render(request, 'alarmasistemas.html', {'form': form})

@login_required
def eliminar_alarmasistemas(request, id):
    AlarmaSistemas.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('alarmasistemaslista'))

@login_required
def editar_alarmasistemas(request, id):
    # se crea el objeto empleado
    alarmasistemas= AlarmaSistemas.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto empleado
    form = AlarmaSistemasForm(instance = alarmasistemas)
    # consulta a todos los empleados (para que se sigan viendo al momento de editar)
    alarmasistemass = AlarmaSistemas.objects.all()
    # se envia el id del empleado para poder marcar con CSS la fila que se esta editando
    return render(request, 'alarmasistemas.html', {'form': form, 'alarmasistemass': alarmasistemass , 'idv': id})

@login_required
def ats(request):
    form = AtsForm()
    ats = Ats.objects.all()
    ctx = {
        'form': form,
        'atss': ats
    }
    return render(request, 'ats.html', ctx)

@login_required
def atslista(request):
    ats = Ats.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            ats = Ats.objects.get(ats=queryset)
        except Exception as e:
            ats = None
   
            ats = Ats.objects.filter(
            Q(codats__contains = queryset ) |
            Q(nombreats__contains = queryset )
             ).distinct()
    ctx = {
        'atss': ats
    }
    return render(request, 'atslista.html', ctx)

@login_required
def guardar_ats(request, id = None):
    if request.method == 'POST':
      ats = get_object_or_404(Ats, pk=id) if id else None
      form = AtsForm(request.POST or None, instance = ats)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('ats'))
    else:
        return render(request, 'ats.html', {'form': form})
#Vista para eliminar ats por su id de identificacion 
@login_required
def eliminar_ats(request, id):
    Ats.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('atslista'))

@login_required
def editar_ats(request, id):
    # se crea el objeto ats
    ats= Ats.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto ats
    form = AtsForm(instance = ats)
    # consulta a todos los ats (para que se sigan viendo al momento de editar)
    atss = Ats.objects.all()
    # se envia el id del ats para poder marcar con CSS la fila que se esta editando
    return render(request, 'ats.html', {'form': form, 'ats': ats , 'idv': id})

#vista para acceder al html de detalle ats
@login_required
def detalleats(request):
    form = DetalleAtsForm()
    detalleats = DetalleAts.objects.all()
    ctx = {
        'form': form,
        'detalleatss': detalleats
    }
    return render(request, 'detalleats.html', ctx)

#vista para poder ver el listado de los detales ats 
#tambien se hace uan consulta para poder filtarr al momento de la busqueda en la barra de busqueda
@login_required
def detalleatslista(request):
    detalleats = DetalleAts.objects.all()
    queryset = request.GET.get('buscar')
    print(queryset)
    if queryset:
        try:
            detalleats = DetalleAts.objects.get(detalleats=queryset)
        except Exception as e:
            detalleats = None
   
            detalleats = DetalleAts.objects.filter(
            Q( lucesencendidasayb__contains = queryset ) |
            Q(preferencia__contains = queryset )
             ).distinct()
    ctx = {
        'detalleatss': detalleats
    }
    return render(request, 'detalleatslista.html', ctx)

#vista para guardar un detalle ats cragando en si el formulario y la consulta para 
#poder insertarla en a base de datos 
@login_required
def guardar_detalleats(request, id = None):
    if request.method == 'POST':
      detalleats = get_object_or_404(DetalleAts, pk=id) if id else None
      form = DetalleAtsForm(request.POST or None, instance = detalleats)
      if form.is_valid():
        form.save()
        return HttpResponseRedirect(reverse('detalleats'))
    else:
        return render(request, 'detalleats.html', {'form': form})

#vista para eliminar un determinado detalle ats adquiriendo su id para ralizarlo
@login_required
def eliminar_detalleats(request, id):
    DetalleAts.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('detalleatslista'))

@login_required
def editar_detalleats(request, id):
    # se crea el objeto detalle ats
    detalleats= DetalleAts.objects.get(pk=id)
    # se crea y se carga el formulario con los datos del objeto  detalle ats
    form = DetalleAtsForm(instance = ats)
    # consulta a todos los  detalle ats (para que se sigan viendo al momento de editar)
    detalleatss = DetalleAts.objects.all()
    # se envia el id del  detalle ats para poder marcar con CSS la fila que se esta editando
    return render(request, 'detalleats.html', {'form': form, 'detalleats': detalleats , 'idv': id})